"""
SCM 물류 브리핑 Excel 자동 생성
─────────────────────────────────────────────────
GitHub Actions에서 매일 자동 실행:
1. 뉴스 요약 데이터 → 뉴스브리핑 시트 (누적)
2. 운임지수 데이터 → 운임지수 시트 (누적)
3. 리스크 추적 → 리스크대장 시트 (자동 분류)

SharePoint에 업로드 가능한 .xlsx 파일로 생성
AI Agent 연계를 위한 정형화된 컬럼 구조
"""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("pipeline/output")
DATA_DIR = Path("data")
BRIEFING_FILE = DATA_DIR / "SCM_물류_브리핑.xlsx"

# 스타일 정의
HEADER_FONT = Font(name="Arial", bold=True, size=10, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
RED_FILL = PatternFill("solid", fgColor="FFF5F5")
YELLOW_FILL = PatternFill("solid", fgColor="FFFBF0")
GREEN_FILL = PatternFill("solid", fgColor="F5FFF5")
DATA_FONT = Font(name="Arial", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# ── 뉴스브리핑 시트 컬럼 ──
NEWS_COLUMNS = [
    ("날짜", 12),
    ("테마", 18),
    ("영향도", 10),
    ("제목", 45),
    ("요약", 55),
    ("영향근거", 40),
    ("대응조치", 35),
    ("출처", 12),
    ("URL", 35),
    ("리스크ID", 12),
]

# ── 운임지수 시트 컬럼 ──
FREIGHT_COLUMNS = [
    ("날짜", 12),
    ("SCFI", 10),
    ("SCFI_기준일", 13),
    ("SCFI_전주대비", 14),
    ("SCFI_YoY", 12),
    ("KCCI", 10),
    ("KCCI_기준일", 13),
    ("KCCI_전주대비", 14),
    ("KCCI_YoY", 12),
]

# ── 리스크대장 시트 컬럼 ──
RISK_COLUMNS = [
    ("리스크ID", 12),
    ("발생일", 12),
    ("테마", 15),
    ("상태", 10),
    ("설명", 45),
    ("영향BL수", 10),
    ("최대지연일", 10),
    ("대응현황", 35),
    ("해소일", 12),
]


def apply_header_style(ws, columns):
    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def get_impact_fill(impact: str):
    if "🔴" in impact:
        return RED_FILL
    elif "🟡" in impact:
        return YELLOW_FILL
    elif "🟢" in impact:
        return GREEN_FILL
    return None


def create_or_load_workbook() -> Workbook:
    if BRIEFING_FILE.exists():
        wb = load_workbook(BRIEFING_FILE)
        return wb

    wb = Workbook()

    # 시트 1: 뉴스브리핑
    ws_news = wb.active
    ws_news.title = "뉴스브리핑"
    apply_header_style(ws_news, NEWS_COLUMNS)

    # 시트 2: 운임지수
    ws_freight = wb.create_sheet("운임지수")
    apply_header_style(ws_freight, FREIGHT_COLUMNS)

    # 시트 3: 리스크대장
    ws_risk = wb.create_sheet("리스크대장")
    apply_header_style(ws_risk, RISK_COLUMNS)

    return wb


def add_news_data(wb: Workbook, summaries: list[dict]):
    ws = wb["뉴스브리핑"]
    today = datetime.now().strftime("%Y-%m-%d")
    risk_counter = get_next_risk_id(wb)

    for s in summaries:
        impact = s.get("impact", "")
        theme = s.get("theme", "")

        # 🔴 뉴스는 자동으로 리스크ID 부여
        risk_id = ""
        if "🔴" in impact:
            risk_id = f"RSK-{risk_counter:03d}"
            risk_counter += 1
            add_risk_entry(wb, risk_id, today, theme, s)

        row = [
            today,
            theme,
            impact,
            s.get("title", ""),
            s.get("summary", ""),
            s.get("impact_reason", ""),
            s.get("action_needed", ""),
            s.get("source", ""),
            s.get("url", ""),
            risk_id,
        ]
        ws.append(row)

        # 행 스타일
        last_row = ws.max_row
        fill = get_impact_fill(impact)
        for col_idx in range(1, len(NEWS_COLUMNS) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if fill:
                cell.fill = fill

    print(f"  📝 뉴스브리핑 {len(summaries)}건 추가")


def get_next_risk_id(wb: Workbook) -> int:
    ws = wb["리스크대장"]
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]).startswith("RSK-"):
            try:
                num = int(str(row[0]).replace("RSK-", ""))
                max_id = max(max_id, num)
            except ValueError:
                pass
    return max_id + 1


def add_risk_entry(wb: Workbook, risk_id: str, date: str, theme: str, news: dict):
    ws = wb["리스크대장"]
    row = [
        risk_id,
        date,
        theme,
        "진행중",
        news.get("title", ""),
        "",   # 영향BL수 (Layer 2 연계 시 자동)
        "",   # 최대지연일
        news.get("action_needed", ""),
        "",   # 해소일
    ]
    ws.append(row)

    last_row = ws.max_row
    for col_idx in range(1, len(RISK_COLUMNS) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.fill = RED_FILL


def add_freight_data(wb: Workbook, latest: dict, history: dict):
    ws = wb["운임지수"]
    today = datetime.now().strftime("%Y-%m-%d")

    scfi = latest.get("scfi") or {}
    kcci = latest.get("kcci") or {}

    # 이미 오늘 데이터가 있는지 확인
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] and str(row[0]) == today:
            print(f"  ℹ️ 운임지수 {today} 이미 존재. 스킵.")
            return

    # 전주 대비 계산
    scfi_cur = scfi.get("current_value")
    scfi_prev = scfi.get("previous_value")
    scfi_wow = ""
    if scfi_cur and scfi_prev:
        diff = scfi_cur - scfi_prev
        pct = (diff / scfi_prev) * 100
        scfi_wow = f"{diff:+,.0f} ({pct:+.1f}%)"

    kcci_cur = kcci.get("current_value")
    kcci_prev = kcci.get("previous_value")
    kcci_wow = ""
    if kcci_cur and kcci_prev:
        diff = kcci_cur - kcci_prev
        pct = (diff / kcci_prev) * 100
        kcci_wow = f"{diff:+,.0f} ({pct:+.1f}%)"

    # YoY 계산
    current_year = datetime.now().year
    prev_year = current_year - 1

    def find_yoy(index_name, current_date):
        if not current_date:
            return ""
        idx_history = history.get(index_name.lower(), [])
        try:
            cur_dt = datetime.strptime(current_date, "%Y-%m-%d")
            cur_week = cur_dt.isocalendar()[1]
            for entry in idx_history:
                if entry.get("date", "").startswith(str(prev_year)):
                    entry_dt = datetime.strptime(entry["date"], "%Y-%m-%d")
                    if abs(entry_dt.isocalendar()[1] - cur_week) <= 2:
                        yoy_val = entry["value"]
                        cur_val = scfi_cur if index_name == "SCFI" else kcci_cur
                        if cur_val and yoy_val:
                            diff = cur_val - yoy_val
                            pct = (diff / yoy_val) * 100
                            return f"{diff:+,.0f} ({pct:+.1f}%)"
        except:
            pass
        return ""

    scfi_yoy = find_yoy("SCFI", scfi.get("current_date"))
    kcci_yoy = find_yoy("KCCI", kcci.get("current_date"))

    row = [today, scfi_cur, scfi.get("current_date", ""), scfi_wow, scfi_yoy,
           kcci_cur, kcci.get("current_date", ""), kcci_wow, kcci_yoy]
    ws.append(row)

    last_row = ws.max_row
    for col_idx in range(1, len(FREIGHT_COLUMNS) + 1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    print(f"  📊 운임지수 추가: SCFI={scfi_cur}, KCCI={kcci_cur}")


def main():
    print("=" * 60)
    print("📋 SCM 물류 브리핑 Excel 생성")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # 데이터 로드
    summary_path = OUTPUT_DIR / "summary.json"
    latest_path = OUTPUT_DIR / "freight_latest.json"
    indices_path = DATA_DIR / "freight_indices.json"

    # Excel 생성 또는 로드
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = create_or_load_workbook()

    # 뉴스 데이터 추가
    if summary_path.exists():
        with open(summary_path, "r", encoding="utf-8") as f:
            summary_data = json.load(f)
        summaries = summary_data.get("summaries", [])
        if summaries:
            add_news_data(wb, summaries)

    # 운임지수 데이터 추가
    if latest_path.exists() and indices_path.exists():
        with open(latest_path, "r", encoding="utf-8") as f:
            latest = json.load(f)
        with open(indices_path, "r", encoding="utf-8") as f:
            history = json.load(f)
        add_freight_data(wb, latest, history)

    # 저장
    wb.save(BRIEFING_FILE)
    print(f"\n💾 Excel 저장 → {BRIEFING_FILE}")

    # output에도 복사 (메일 첨부용)
    output_copy = OUTPUT_DIR / "SCM_물류_브리핑.xlsx"
    wb.save(output_copy)
    print(f"💾 메일 첨부용 복사 → {output_copy}")


if __name__ == "__main__":
    main()
